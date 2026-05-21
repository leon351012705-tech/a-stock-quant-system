"""
run_save_stats.py
每日扫描结果统计 — 自动写入Excel，每月一个Sheet

用法：在 run_signal_scan.py 跑完后调用，或单独运行
统计文件保存在项目根目录：signal_stats.xlsx
"""

import os
import sys
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import DB_PATH

STATS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "signal_stats.xlsx")

# 列定义
HEADERS = [
    "日期", "市场状态", "上涨占比%", "中位涨跌%",
    "共振信号数", "共振股票", "共振策略", "共振触发日",
    "MACD命中数", "BOLL命中数", "MA命中数", "SSB命中数",
    "备注"
]

COL_WIDTHS = [12, 10, 10, 10, 10, 30, 20, 25, 10, 10, 10, 10, 20]


def get_today_stats() -> dict:
    """从数据库读取今日扫描数据"""
    conn = sqlite3.connect(DB_PATH)

    # 最新交易日
    latest_date = pd.read_sql(
        "SELECT MAX(trade_date) as d FROM daily_bars", conn
    ).iloc[0]["d"]

    # 市场广度
    breadth = pd.read_sql(
        "SELECT pct_change FROM daily_bars WHERE trade_date = ? AND pct_change IS NOT NULL",
        conn, params=(latest_date,),
    )

    if len(breadth) >= 100:
        total = len(breadth)
        up_ratio = round((breadth["pct_change"] > 0).sum() / total * 100, 1)
        median_pct = round(breadth["pct_change"].median(), 2)
        big_drop = (breadth["pct_change"] < -4).sum() / total

        # 判断市场状态
        day_ok = (
            (up_ratio/100 >= 0.45 and big_drop <= 0.20)
            or (up_ratio/100 >= 0.40 and median_pct >= -0.5)
        )
        market_status = "✅可做多" if day_ok else "❌不可做"
    else:
        up_ratio = None
        median_pct = None
        market_status = "数据不足"

    # 各策略命中数（从缓存读）
    cache = pd.read_sql(
        """
        SELECT strategy_id, COUNT(DISTINCT symbol) as cnt
        FROM signal_cache WHERE trade_date = ?
        GROUP BY strategy_id
        """,
        conn, params=(latest_date,),
    )

    strategy_counts = dict(zip(cache["strategy_id"], cache["cnt"])) if not cache.empty else {}

    # 共振信号（从signal_cache重建）
    resonance_rows = []
    if not cache.empty:
        from collections import defaultdict

        TREND = {"macd", "ma_trend"}

        sym_strats = defaultdict(set)
        sym_dates  = defaultdict(set)

        # 取最近3个交易日
        recent = pd.read_sql(
            """
            SELECT DISTINCT trade_date FROM daily_bars
            WHERE trade_date <= ? ORDER BY trade_date DESC LIMIT 3
            """,
            conn, params=(latest_date,),
        )["trade_date"].tolist()

        for td in recent:
            day = pd.read_sql(
                "SELECT strategy_id, symbol FROM signal_cache WHERE trade_date = ?",
                conn, params=(td,),
            )
            for _, row in day.iterrows():
                sym_strats[row["symbol"]].add(row["strategy_id"])
                sym_dates[row["symbol"]].add(td)

        # 一次取所有候选股的名字（旧实现：每只 1 次查询）
        candidates = [s for s, strats in sym_strats.items()
                      if "boll_rv" in strats and bool(strats & TREND)]
        names = {}
        if candidates:
            placeholders = ",".join(["?"] * len(candidates))
            name_df = pd.read_sql(
                f"SELECT symbol, name FROM stock_info WHERE symbol IN ({placeholders})",
                conn, params=candidates,
            )
            names = dict(zip(name_df["symbol"], name_df["name"]))

        for sym in candidates:
            resonance_rows.append({
                "symbol": sym,
                "name":   names.get(sym, sym),
                "strats": ",".join(sorted(sym_strats[sym])),
                "dates":  ",".join(sorted(sym_dates[sym])),
            })

    conn.close()

    return {
        "date":           latest_date,
        "market_status":  market_status,
        "up_ratio":       up_ratio,
        "median_pct":     median_pct,
        "resonance":      resonance_rows,
        "macd_count":     strategy_counts.get("macd", 0),
        "boll_count":     strategy_counts.get("boll_rv", 0),
        "ma_count":       strategy_counts.get("ma_trend", 0),
        "ssb_count":      strategy_counts.get("ssb", 0),
    }


def get_or_create_workbook():
    if os.path.exists(STATS_FILE):
        return load_workbook(STATS_FILE)
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认空sheet
    return wb


def get_or_create_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # 写表头
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="2F5597")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
        )
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    return ws


def row_exists(ws, date_str: str) -> bool:
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == date_str:
            return True
    return False


def write_row(ws, stats: dict):
    resonance = stats["resonance"]
    res_count = len(resonance)
    res_stocks = "  ".join([f"{r['name']}({r['symbol']})" for r in resonance]) if resonance else "无"
    res_strats = "  ".join([r["strats"] for r in resonance]) if resonance else "-"
    res_dates  = "  ".join([r["dates"]  for r in resonance]) if resonance else "-"

    row_data = [
        stats["date"],
        stats["market_status"],
        stats["up_ratio"],
        stats["median_pct"],
        res_count,
        res_stocks,
        res_strats,
        res_dates,
        stats["macd_count"],
        stats["boll_count"],
        stats["ma_count"],
        stats["ssb_count"],
        "",  # 备注留空手填
    ]

    next_row = ws.max_row + 1
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 交替行背景
        if next_row % 2 == 0:
            cell.fill = PatternFill("solid", start_color="EEF2FF")

        # 市场状态列颜色
        if col == 2:
            if "可做多" in str(value):
                cell.font = Font(name="Arial", size=9, color="1F7A3A", bold=True)
            else:
                cell.font = Font(name="Arial", size=9, color="CC0000", bold=True)

        # 共振信号数列颜色
        if col == 5 and isinstance(value, int) and value > 0:
            cell.font = Font(name="Arial", size=9, color="CC6600", bold=True)

    ws.row_dimensions[next_row].height = 30


def main():
    print("读取今日扫描数据...")
    stats = get_today_stats()

    date_str = stats["date"]
    sheet_name = date_str[:7]  # "2026-04"

    print(f"交易日：{date_str}  Sheet：{sheet_name}")
    print(f"市场：{stats['market_status']}  共振信号：{len(stats['resonance'])}只")

    wb = get_or_create_workbook()
    ws = get_or_create_sheet(wb, sheet_name)

    if row_exists(ws, date_str):
        print(f"⚠️  {date_str} 数据已存在，跳过写入")
    else:
        write_row(ws, stats)
        print(f"✅ 已写入 {date_str} 数据")

    wb.save(STATS_FILE)
    print(f"统计文件已保存：{STATS_FILE}")


if __name__ == "__main__":
    main()
